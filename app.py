"""
Igreja CEADREI — servidor Flask com páginas HTML, API JSON e galeria admin.
"""

from __future__ import annotations

import io
import os
import uuid
from functools import wraps
from pathlib import Path

from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

import arraial
import batismo
import casais
import gallery
import mocidade
import pastores

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "ceadrei-dev-secret-change-me")
app.config["MAX_CONTENT_LENGTH"] = 12 * 1024 * 1024  # 12 MB por upload

# Senha do painel da mídia (troque em produção via variável de ambiente)
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "ceasdrei")
ADMIN_PASSWORD_HASH = generate_password_hash(ADMIN_PASSWORD)

# Senha do painel do Evento Batismo (Evangelista Sueli)
BATISMO_PASSWORD = os.environ.get("BATISMO_PASSWORD", "Sueli")
BATISMO_PASSWORD_HASH = generate_password_hash(BATISMO_PASSWORD)

# Senha do painel do Encontro de Casais (Diac. Robson & Diac. Luana)
CASAIS_PASSWORD = os.environ.get("CASAIS_PASSWORD", "Robson&Luana")
CASAIS_PASSWORD_HASH = generate_password_hash(CASAIS_PASSWORD)

# Senha do painel dos Pastores
PASTORES_PASSWORD = os.environ.get("PASTORES_PASSWORD", "Solange123")
PASTORES_PASSWORD_HASH = generate_password_hash(PASTORES_PASSWORD)

# Senha do departamento Arraiá / Cantina (Diac. Cássia)
ARRAIAL_PASSWORD = os.environ.get("ARRAIAL_PASSWORD", "Cassia")
ARRAIAL_PASSWORD_HASH = generate_password_hash(ARRAIAL_PASSWORD)

# Senha dos Líderes da Mocidade (Diác. Natan e Diác. Ana Beatriz)
MOCIDADE_PASSWORD = os.environ.get("MOCIDADE_PASSWORD", "Mocidade")
MOCIDADE_PASSWORD_HASH = generate_password_hash(MOCIDADE_PASSWORD)


def load_json(name: str) -> dict:
    path = DATA_DIR / name
    if not path.exists():
        return {}
    return __import__("json").loads(path.read_text(encoding="utf-8"))


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_ok"):
            return redirect(url_for("admin_login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def batismo_login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("batismo_ok"):
            return redirect(url_for("batismo_login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def casais_login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("casais_ok"):
            return redirect(url_for("casais_login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def pastores_login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("pastores_ok"):
            return redirect(url_for("pastores_login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def arraial_login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("arraial_ok"):
            return redirect(url_for("arraial_login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def mocidade_login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("mocidade_ok"):
            return redirect(url_for("mocidade_login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


@app.context_processor
def inject_admin():
    return {
        "admin_logado": bool(session.get("admin_ok")),
        "batismo_logado": bool(session.get("batismo_ok")),
        "casais_logado": bool(session.get("casais_ok")),
        "pastores_logado": bool(session.get("pastores_ok")),
        "arraial_logado": bool(session.get("arraial_ok")),
        "mocidade_logado": bool(session.get("mocidade_ok")),
    }


@app.route("/")
def home():
    igreja = load_json("igreja.json")
    cultos = load_json("cultos.json").get("cultos", [])[:3]
    posts = gallery.listar_posts_ativos()
    destaque = posts[0] if posts else None
    fotos_batismo = batismo.listar_fotos()
    fotos_casais = casais.listar_fotos()
    programacao_casais = casais.obter_programacao()
    escala_hoje = pastores.obter_proxima_escala()
    destaque_culto = pastores.obter_destaque()
    avisos_lideres = pastores.avisos_proximos(7)[:3]
    info_arraial = arraial.info_evento()
    post_mocidade = mocidade.obter_post_ativo()
    return render_template(
        "index.html",
        igreja=igreja,
        cultos=cultos,
        galeria_destaque=destaque,
        batismo_fotos=fotos_batismo,
        casais_fotos=fotos_casais,
        casais_programacao=programacao_casais,
        escala_hoje=escala_hoje,
        destaque_culto=destaque_culto,
        avisos_lideres=avisos_lideres,
        arraial=info_arraial,
        cantina_texto=arraial.obter_cantina(),
        post_mocidade=post_mocidade,
    )


@app.route("/sobre")
def sobre():
    igreja = load_json("igreja.json")
    return render_template("sobre.html", igreja=igreja)


@app.route("/cultos")
def cultos_page():
    igreja = load_json("igreja.json")
    cultos = load_json("cultos.json").get("cultos", [])
    return render_template("cultos.html", igreja=igreja, cultos=cultos)


@app.route("/eventos")
def eventos_page():
    igreja = load_json("igreja.json")
    eventos = load_json("eventos.json").get("eventos", [])
    return render_template("eventos.html", igreja=igreja, eventos=eventos)


@app.route("/contato")
def contato():
    igreja = load_json("igreja.json")
    return render_template("contato.html", igreja=igreja)


@app.route("/galeria")
def galeria_publica():
    igreja = load_json("igreja.json")
    posts = gallery.listar_posts_ativos()
    return render_template("galeria.html", igreja=igreja, posts=posts)


# ---------- Admin (mídia da igreja) ----------

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    igreja = load_json("igreja.json")
    erro = None
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if check_password_hash(ADMIN_PASSWORD_HASH, senha):
            session["admin_ok"] = True
            destino = request.args.get("next") or url_for("admin_galeria")
            return redirect(destino)
        erro = "Senha incorreta. Tente novamente."
    if session.get("admin_ok"):
        return redirect(url_for("admin_galeria"))
    return render_template("admin_login.html", igreja=igreja, erro=erro)


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_ok", None)
    return redirect(url_for("home"))


@app.route("/admin/galeria", methods=["GET", "POST"])
@login_required
def admin_galeria():
    igreja = load_json("igreja.json")
    cultos = load_json("cultos.json").get("cultos", [])

    if request.method == "POST":
        culto_titulo = request.form.get("culto_titulo", "").strip()
        culto_dia = request.form.get("culto_dia", "").strip()
        titulo = request.form.get("titulo", "").strip() or f"Fotos — {culto_titulo}"
        arquivos = request.files.getlist("fotos")

        salvos = []
        gallery.init_db()
        for arquivo in arquivos:
            if not arquivo or not arquivo.filename:
                continue
            if not gallery.extensao_ok(arquivo.filename):
                flash(f"Formato não permitido: {arquivo.filename}", "erro")
                continue
            nome_seguro = secure_filename(arquivo.filename)
            extensao = Path(nome_seguro).suffix.lower()
            nome_final = f"{uuid.uuid4().hex}{extensao}"
            destino = gallery.UPLOAD_DIR / nome_final
            arquivo.save(destino)
            salvos.append(nome_final)

        if not culto_titulo or not culto_dia:
            flash("Escolha o culto e o dia.", "erro")
        elif not salvos:
            flash("Envie ao menos uma foto válida (JPG, PNG, WEBP ou GIF).", "erro")
        else:
            gallery.criar_post(culto_titulo, culto_dia, titulo, salvos)
            flash("Postagem publicada! As fotos ficam no ar até você apagar no painel.", "ok")
            return redirect(url_for("admin_galeria"))

    posts = gallery.listar_posts_ativos()
    return render_template(
        "admin_galeria.html",
        igreja=igreja,
        cultos=cultos,
        posts=posts,
    )


@app.route("/admin/galeria/<int:post_id>/apagar", methods=["POST"])
@login_required
def admin_apagar_post(post_id: int):
    if gallery.apagar_post(post_id):
        flash("Postagem apagada.", "ok")
    else:
        flash("Postagem não encontrada.", "erro")
    return redirect(url_for("admin_galeria"))


# ---------- Evento Batismo ----------

@app.route("/batismo/login", methods=["GET", "POST"])
def batismo_login():
    igreja = load_json("igreja.json")
    erro = None
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if check_password_hash(BATISMO_PASSWORD_HASH, senha):
            session["batismo_ok"] = True
            destino = request.args.get("next") or url_for("batismo_admin")
            return redirect(destino)
        erro = "Senha incorreta. Tente novamente."
    if session.get("batismo_ok"):
        return redirect(url_for("batismo_admin"))
    return render_template("batismo_login.html", igreja=igreja, erro=erro)


@app.route("/batismo/logout")
def batismo_logout():
    session.pop("batismo_ok", None)
    return redirect(url_for("eventos_page"))


def _processar_evento_responsavel(origem: str) -> bool:
    """Salva evento do responsável no calendário compartilhado. Retorna True se processou."""
    info = pastores.RESPONSAVEIS_EVENTO.get(origem)
    if not info:
        return False
    data_iso = request.form.get("data", "").strip()
    if not data_iso:
        flash("Informe a data do seu evento.", "erro")
        return True
    evento_id = request.form.get("evento_id", type=int)
    titulo = request.form.get("titulo", "").strip() or info["titulo"]
    pastores.salvar_evento_lider(
        titulo=titulo,
        lider=info["lider"],
        origem=origem,
        data_iso=data_iso,
        horario=request.form.get("horario", ""),
        local=request.form.get("local", ""),
        descricao=request.form.get("descricao", ""),
        aviso=request.form.get("aviso", ""),
        evento_id=evento_id,
    )
    flash("Evento registrado no calendário dos líderes.", "ok")
    return True


@app.route("/batismo/admin", methods=["GET", "POST"])
@batismo_login_required
def batismo_admin():
    igreja = load_json("igreja.json")

    if request.method == "POST":
        acao = request.form.get("acao", "fotos").strip()
        if acao == "evento":
            _processar_evento_responsavel("batismo")
            return redirect(url_for("batismo_admin"))

        titulo = request.form.get("titulo", "").strip() or "Local do batismo"
        arquivos = request.files.getlist("fotos")
        salvos = []
        batismo.init_db()
        for arquivo in arquivos:
            if not arquivo or not arquivo.filename:
                continue
            if not batismo.extensao_ok(arquivo.filename):
                flash(f"Formato não permitido: {arquivo.filename}", "erro")
                continue
            nome_seguro = secure_filename(arquivo.filename)
            extensao = Path(nome_seguro).suffix.lower()
            nome_final = f"{uuid.uuid4().hex}{extensao}"
            destino = batismo.UPLOAD_DIR / nome_final
            arquivo.save(destino)
            salvos.append(nome_final)

        if not salvos:
            flash("Envie ao menos uma foto válida (JPG, PNG, WEBP ou GIF).", "erro")
        else:
            batismo.adicionar_fotos(salvos, titulo)
            flash("Fotos do sítio publicadas! Elas aparecem na página inicial.", "ok")
            return redirect(url_for("batismo_admin"))

    editar_evento = None
    if request.args.get("editar_evento"):
        editar_evento = pastores.obter_evento_lider(
            request.args.get("editar_evento", type=int),
            origem="batismo",
        )

    return render_template(
        "batismo_admin.html",
        igreja=igreja,
        fotos=batismo.listar_fotos(),
        inscricoes=batismo.listar_inscricoes(),
        status_opcoes=batismo.STATUS_OPCOES,
        eventos_calendario=pastores.listar_eventos_lideres(origem="batismo"),
        editar_evento=editar_evento,
        info_evento=pastores.RESPONSAVEIS_EVENTO["batismo"],
    )


@app.route("/batismo/admin/foto/<int:foto_id>/apagar", methods=["POST"])
@batismo_login_required
def batismo_apagar_foto(foto_id: int):
    if batismo.apagar_foto(foto_id):
        flash("Foto apagada.", "ok")
    else:
        flash("Foto não encontrada.", "erro")
    return redirect(url_for("batismo_admin"))


@app.route("/batismo/admin/inscricao/<int:inscricao_id>/status", methods=["POST"])
@batismo_login_required
def batismo_atualizar_status(inscricao_id: int):
    status = request.form.get("status", "").strip()
    if batismo.atualizar_status(inscricao_id, status):
        flash("Status atualizado.", "ok")
    else:
        flash("Não foi possível atualizar o status.", "erro")
    return redirect(url_for("batismo_admin"))


@app.route("/batismo/admin/evento/<int:evento_id>/apagar", methods=["POST"])
@batismo_login_required
def batismo_apagar_evento(evento_id: int):
    if pastores.apagar_evento_lider(evento_id, origem="batismo"):
        flash("Evento removido do calendário.", "ok")
    else:
        flash("Evento não encontrado.", "erro")
    return redirect(url_for("batismo_admin"))


@app.route("/batismo/admin/inscricao/<int:inscricao_id>/apagar", methods=["POST"])
@batismo_login_required
def batismo_apagar_inscricao(inscricao_id: int):
    if batismo.apagar_inscricao(inscricao_id):
        flash("Família removida da lista de batismo.", "ok")
    else:
        flash("Inscrição não encontrada.", "erro")
    return redirect(url_for("batismo_admin"))


@app.route("/batismo/inscricao", methods=["GET", "POST"])
def batismo_inscricao():
    igreja = load_json("igreja.json")
    fotos = batismo.listar_fotos()
    erro = None

    if request.method == "POST":
        nome_marido = request.form.get("nome_marido", "").strip()
        nome_mulher = request.form.get("nome_mulher", "").strip()
        rg_marido = request.form.get("rg_marido", "").strip()
        rg_mulher = request.form.get("rg_mulher", "").strip()
        nomes_filhos = request.form.getlist("filhos")
        docs_filhos = request.form.getlist("filhos_doc")
        filhos = []
        for i, nome in enumerate(nomes_filhos):
            nome_limpo = (nome or "").strip()
            if not nome_limpo:
                continue
            doc = docs_filhos[i].strip() if i < len(docs_filhos) else ""
            filhos.append({"nome": nome_limpo, "documento": doc})
        telefone = request.form.get("telefone", "").strip()
        participantes = request.form.getlist("participantes")
        status = request.form.get("status", "analise").strip()

        # Identificação da família (sem campo "nome completo")
        if nome_marido and nome_mulher:
            nome_completo = f"{nome_marido} & {nome_mulher}"
        else:
            nome_completo = nome_marido or nome_mulher or "Família"

        if not nome_marido and not nome_mulher:
            erro = "Informe o nome do marido e/ou da mulher."
        elif not telefone:
            erro = "Informe o telefone de contato."
        elif not participantes:
            erro = "Selecione quem vai ao batismo (marido, mulher e/ou filhos)."
        elif status not in batismo.STATUS_OPCOES:
            erro = "Selecione uma confirmação válida."
        else:
            inscricao_id = batismo.criar_inscricao(
                nome_completo=nome_completo,
                nome_marido=nome_marido,
                nome_mulher=nome_mulher,
                filhos=filhos,
                rg_marido=rg_marido,
                rg_mulher=rg_mulher,
                telefone=telefone,
                participantes=participantes,
                status=status,
            )
            return redirect(url_for("batismo_confirmacao", inscricao_id=inscricao_id))

    return render_template(
        "batismo_inscricao.html",
        igreja=igreja,
        fotos=fotos,
        erro=erro,
        status_opcoes=batismo.STATUS_OPCOES,
        participante_opcoes=batismo.PARTICIPANTE_OPCOES,
    )


@app.route("/batismo/confirmacao/<int:inscricao_id>")
def batismo_confirmacao(inscricao_id: int):
    igreja = load_json("igreja.json")
    inscricao = batismo.obter_inscricao(inscricao_id)
    if not inscricao:
        flash("Inscrição não encontrada.", "erro")
        return redirect(url_for("batismo_inscricao"))
    return render_template(
        "batismo_confirmacao.html",
        igreja=igreja,
        inscricao=inscricao,
    )


@app.route("/batismo/admin/exportar.xlsx")
@batismo_login_required
def batismo_exportar_excel():
    from openpyxl import Workbook

    inscricoes = batismo.listar_inscricoes()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inscricoes Batismo"
    ws.append(
        [
            "ID",
            "Marido",
            "RG/CPF Marido",
            "Mulher",
            "RG/CPF Mulher",
            "Filhos",
            "Telefone",
            "Participantes",
            "Status",
            "Enviado em",
        ]
    )
    for item in inscricoes:
        ws.append(
            [
                item["id"],
                item["nome_marido"],
                item.get("rg_marido") or "",
                item["nome_mulher"],
                item.get("rg_mulher") or "",
                item.get("filhos_texto") or "",
                item["telefone"],
                item["participantes_texto"],
                item["status_texto"],
                item["criado_em"].replace("T", " "),
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inscricoes-batismo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/batismo/admin/exportar.pdf")
@batismo_login_required
def batismo_exportar_pdf():
    from fpdf import FPDF

    inscricoes = batismo.listar_inscricoes()
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Inscricoes - Evento Batismo CEADREI", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(
        0,
        8,
        "Responsavel: Evangelista Sueli",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    colunas = [
        ("ID", 12),
        ("Nome", 45),
        ("Marido", 35),
        ("Mulher", 35),
        ("Filhos", 40),
        ("Telefone", 30),
        ("Quem", 30),
        ("Status", 35),
    ]
    pdf.set_font("Helvetica", "B", 8)
    for titulo, largura in colunas:
        pdf.cell(largura, 8, titulo, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for item in inscricoes:
        valores = [
            str(item["id"]),
            item["nome_completo"][:40],
            item["nome_marido"][:30],
            item["nome_mulher"][:30],
            (item.get("filhos_texto") or item.get("filhos") or "")[:35],
            item["telefone"][:22],
            item["participantes_texto"][:25],
            item["status_texto"][:28],
        ]
        for valor, (_, largura) in zip(valores, colunas):
            pdf.cell(largura, 7, valor, border=1)
        pdf.ln()

    if not inscricoes:
        pdf.cell(0, 10, "Nenhuma inscricao registrada.", new_x="LMARGIN", new_y="NEXT")

    buffer = io.BytesIO(pdf.output())
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inscricoes-batismo.pdf",
        mimetype="application/pdf",
    )


# ---------- Encontro de Casais ----------

@app.route("/casais/login", methods=["GET", "POST"])
def casais_login():
    igreja = load_json("igreja.json")
    erro = None
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if check_password_hash(CASAIS_PASSWORD_HASH, senha):
            session["casais_ok"] = True
            destino = request.args.get("next") or url_for("casais_admin")
            return redirect(destino)
        erro = "Senha incorreta. Tente novamente."
    if session.get("casais_ok"):
        return redirect(url_for("casais_admin"))
    return render_template("casais_login.html", igreja=igreja, erro=erro)


@app.route("/casais/logout")
def casais_logout():
    session.pop("casais_ok", None)
    return redirect(url_for("eventos_page"))


@app.route("/casais/admin", methods=["GET", "POST"])
@casais_login_required
def casais_admin():
    igreja = load_json("igreja.json")

    if request.method == "POST":
        acao = request.form.get("acao", "fotos").strip()

        if acao == "programacao":
            casais.salvar_programacao(request.form.get("programacao", ""))
            flash("Programação do encontro atualizada.", "ok")
            return redirect(url_for("casais_admin"))

        if acao == "evento":
            _processar_evento_responsavel("casais")
            return redirect(url_for("casais_admin"))

        titulo = request.form.get("titulo", "").strip() or "Encontro de Casais"
        arquivos = request.files.getlist("fotos")
        salvos = []
        casais.init_db()
        for arquivo in arquivos:
            if not arquivo or not arquivo.filename:
                continue
            if not casais.extensao_ok(arquivo.filename):
                flash(f"Formato não permitido: {arquivo.filename}", "erro")
                continue
            nome_seguro = secure_filename(arquivo.filename)
            extensao = Path(nome_seguro).suffix.lower()
            nome_final = f"{uuid.uuid4().hex}{extensao}"
            destino = casais.UPLOAD_DIR / nome_final
            arquivo.save(destino)
            salvos.append(nome_final)

        if not salvos:
            flash("Envie ao menos uma foto válida (JPG, PNG, WEBP ou GIF).", "erro")
        else:
            casais.adicionar_fotos(salvos, titulo)
            flash("Fotos publicadas! Elas aparecem na página do encontro e no início.", "ok")
            return redirect(url_for("casais_admin"))

    editar_evento = None
    if request.args.get("editar_evento"):
        editar_evento = pastores.obter_evento_lider(
            request.args.get("editar_evento", type=int),
            origem="casais",
        )

    return render_template(
        "casais_admin.html",
        igreja=igreja,
        fotos=casais.listar_fotos(),
        inscricoes=casais.listar_inscricoes(),
        programacao=casais.obter_programacao(),
        status_opcoes=casais.STATUS_OPCOES,
        h1_responsaveis=casais.H1_RESPONSAVEIS,
        eventos_calendario=pastores.listar_eventos_lideres(origem="casais"),
        editar_evento=editar_evento,
        info_evento=pastores.RESPONSAVEIS_EVENTO["casais"],
    )


@app.route("/casais/admin/foto/<int:foto_id>/apagar", methods=["POST"])
@casais_login_required
def casais_apagar_foto(foto_id: int):
    if casais.apagar_foto(foto_id):
        flash("Foto apagada.", "ok")
    else:
        flash("Foto não encontrada.", "erro")
    return redirect(url_for("casais_admin"))


@app.route("/casais/admin/inscricao/<int:inscricao_id>/status", methods=["POST"])
@casais_login_required
def casais_atualizar_status(inscricao_id: int):
    status = request.form.get("status", "").strip()
    if casais.atualizar_status(inscricao_id, status):
        flash("Status atualizado.", "ok")
    else:
        flash("Não foi possível atualizar o status.", "erro")
    return redirect(url_for("casais_admin"))


@app.route("/casais/admin/evento/<int:evento_id>/apagar", methods=["POST"])
@casais_login_required
def casais_apagar_evento(evento_id: int):
    if pastores.apagar_evento_lider(evento_id, origem="casais"):
        flash("Evento removido do calendário.", "ok")
    else:
        flash("Evento não encontrado.", "erro")
    return redirect(url_for("casais_admin"))


@app.route("/casais")
def casais_page():
    igreja = load_json("igreja.json")
    return render_template(
        "casais.html",
        igreja=igreja,
        fotos=casais.listar_fotos(),
        programacao=casais.obter_programacao(),
        h1_responsaveis=casais.H1_RESPONSAVEIS,
    )


@app.route("/casais/inscricao", methods=["GET", "POST"])
def casais_inscricao():
    igreja = load_json("igreja.json")
    erro = None

    if request.method == "POST":
        nome_marido = request.form.get("nome_marido", "").strip()
        telefone_marido = request.form.get("telefone_marido", "").strip()
        nome_mulher = request.form.get("nome_mulher", "").strip()
        telefone_mulher = request.form.get("telefone_mulher", "").strip()
        status = request.form.get("status", "analise").strip()

        if not nome_marido or not nome_mulher:
            erro = "Informe o nome do marido e da mulher."
        elif not telefone_marido and not telefone_mulher:
            erro = "Informe ao menos um telefone de contato."
        elif status not in casais.STATUS_OPCOES:
            erro = "Selecione uma confirmação válida."
        else:
            inscricao_id = casais.criar_inscricao(
                nome_marido=nome_marido,
                telefone_marido=telefone_marido,
                nome_mulher=nome_mulher,
                telefone_mulher=telefone_mulher,
                status=status,
            )
            return redirect(url_for("casais_confirmacao", inscricao_id=inscricao_id))

    return render_template(
        "casais_inscricao.html",
        igreja=igreja,
        fotos=casais.listar_fotos(),
        programacao=casais.obter_programacao(),
        erro=erro,
        status_opcoes=casais.STATUS_OPCOES,
        h1_responsaveis=casais.H1_RESPONSAVEIS,
    )


@app.route("/casais/confirmacao/<int:inscricao_id>")
def casais_confirmacao(inscricao_id: int):
    igreja = load_json("igreja.json")
    inscricao = casais.obter_inscricao(inscricao_id)
    if not inscricao:
        flash("Inscrição não encontrada.", "erro")
        return redirect(url_for("casais_inscricao"))
    return render_template(
        "casais_confirmacao.html",
        igreja=igreja,
        inscricao=inscricao,
        h1_responsaveis=casais.H1_RESPONSAVEIS,
    )


@app.route("/casais/admin/exportar.xlsx")
@casais_login_required
def casais_exportar_excel():
    from openpyxl import Workbook

    inscricoes = casais.listar_inscricoes()
    wb = Workbook()
    ws = wb.active
    ws.title = "Encontro Casais"
    ws.append(
        [
            "ID",
            "Marido",
            "Telefone marido",
            "Mulher",
            "Telefone mulher",
            "Status",
            "Enviado em",
        ]
    )
    for item in inscricoes:
        ws.append(
            [
                item["id"],
                item["nome_marido"],
                item["telefone_marido"],
                item["nome_mulher"],
                item["telefone_mulher"],
                item["status_texto"],
                item["criado_em"].replace("T", " "),
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inscricoes-encontro-casais.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/casais/admin/exportar.pdf")
@casais_login_required
def casais_exportar_pdf():
    from fpdf import FPDF

    inscricoes = casais.listar_inscricoes()
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Inscricoes - Encontro de Casais CEADREI", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(
        0,
        8,
        "Responsaveis: Diac. Robson e Diac. Luana",
        new_x="LMARGIN",
        new_y="NEXT",
    )
    pdf.ln(2)

    colunas = [
        ("ID", 12),
        ("Marido", 50),
        ("Tel. marido", 35),
        ("Mulher", 50),
        ("Tel. mulher", 35),
        ("Status", 40),
        ("Enviado em", 40),
    ]
    pdf.set_font("Helvetica", "B", 8)
    for titulo, largura in colunas:
        pdf.cell(largura, 8, titulo, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    for item in inscricoes:
        valores = [
            str(item["id"]),
            item["nome_marido"][:40],
            item["telefone_marido"][:24],
            item["nome_mulher"][:40],
            item["telefone_mulher"][:24],
            item["status_texto"][:30],
            item["criado_em"].replace("T", " ")[:22],
        ]
        for valor, (_, largura) in zip(valores, colunas):
            pdf.cell(largura, 7, valor, border=1)
        pdf.ln()

    if not inscricoes:
        pdf.cell(0, 10, "Nenhuma inscricao registrada.", new_x="LMARGIN", new_y="NEXT")

    buffer = io.BytesIO(pdf.output())
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="inscricoes-encontro-casais.pdf",
        mimetype="application/pdf",
    )


# ---------- Pastores / Obreiros / Calendário ----------

@app.route("/pastores/login", methods=["GET", "POST"])
def pastores_login():
    igreja = load_json("igreja.json")
    erro = None
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if check_password_hash(PASTORES_PASSWORD_HASH, senha):
            session["pastores_ok"] = True
            destino = request.args.get("next") or url_for("pastores_admin")
            return redirect(destino)
        erro = "Senha incorreta. Tente novamente."
    if session.get("pastores_ok"):
        return redirect(url_for("pastores_admin"))
    return render_template("pastores_login.html", igreja=igreja, erro=erro)


@app.route("/pastores/logout")
def pastores_logout():
    session.pop("pastores_ok", None)
    return redirect(url_for("eventos_page"))


@app.route("/pastores/admin", methods=["GET", "POST"])
@pastores_login_required
def pastores_admin():
    igreja = load_json("igreja.json")
    hoje = pastores.hoje()
    ano = int(request.args.get("ano", hoje.year))
    mes = int(request.args.get("mes", hoje.month))
    if mes < 1 or mes > 12:
        mes = hoje.month

    if request.method == "POST":
        acao = request.form.get("acao", "").strip()

        if acao == "escala":
            data_iso = request.form.get("data", "").strip()
            if not data_iso:
                flash("Informe a data da escala.", "erro")
            else:
                escala_id = request.form.get("escala_id", type=int)
                pastores.salvar_escala(
                    data_iso=data_iso,
                    porta_vidro=request.form.get("porta_vidro", ""),
                    abertura=request.form.get("abertura", ""),
                    porta_escada=request.form.get("porta_escada", ""),
                    escala_id=escala_id,
                )
                flash("Escala de obreiros salva.", "ok")
            return redirect(url_for("pastores_admin", ano=ano, mes=mes, aba="escala"))

        if acao == "destaque":
            pregador_nome = request.form.get("pregador_nome", "").strip()
            arquivo = request.files.get("pregador_foto")
            foto_nome = ""
            if pregador_nome:
                if arquivo and arquivo.filename:
                    if not pastores.extensao_ok(arquivo.filename):
                        flash("Foto do pregador: use JPG, PNG, WEBP ou GIF.", "erro")
                        return redirect(url_for("pastores_admin", aba="destaque"))
                    nome_seguro = secure_filename(arquivo.filename)
                    extensao = Path(nome_seguro).suffix.lower()
                    foto_nome = f"pregador-{uuid.uuid4().hex}{extensao}"
                    arquivo.save(pastores.UPLOAD_DIR / foto_nome)
                elif not pastores.obter_destaque().get("pregador_foto"):
                    flash("Ao informar o pregador, envie também a foto.", "erro")
                    return redirect(url_for("pastores_admin", aba="destaque"))

            pastores.salvar_destaque(
                data_iso=request.form.get("data", "").strip(),
                porta=request.form.get("porta", ""),
                abertura_culto=request.form.get("abertura_culto", ""),
                pregador_nome=pregador_nome,
                pregador_foto=foto_nome,
                mensagem_campanha=request.form.get("mensagem_campanha", ""),
                referencia=request.form.get("referencia", ""),
                manter_foto=True,
            )
            flash("Destaque do culto atualizado.", "ok")
            return redirect(url_for("pastores_admin", aba="destaque"))

    editar_escala = None
    if request.args.get("editar_escala"):
        for item in pastores.listar_escala(ano, mes):
            if item["id"] == request.args.get("editar_escala", type=int):
                editar_escala = item
                break

    return render_template(
        "pastores_admin.html",
        igreja=igreja,
        h1_pastores=pastores.H1_PASTORES,
        escala=pastores.listar_escala(ano, mes),
        destaque=pastores.obter_destaque(),
        eventos=pastores.listar_eventos_lideres(incluir_passados=True),
        calendario=pastores.calendario_mes(ano, mes),
        avisos=pastores.avisos_proximos(7),
        editar_escala=editar_escala,
        ano=ano,
        mes=mes,
        aba=request.args.get("aba", "escala"),
    )


@app.route("/pastores/admin/escala/<int:escala_id>/apagar", methods=["POST"])
@pastores_login_required
def pastores_apagar_escala(escala_id: int):
    if pastores.apagar_escala(escala_id):
        flash("Dia removido da escala.", "ok")
    else:
        flash("Registro não encontrado.", "erro")
    return redirect(url_for("pastores_admin", aba="escala"))


@app.route("/pastores/admin/evento/<int:evento_id>/apagar", methods=["POST"])
@pastores_login_required
def pastores_apagar_evento(evento_id: int):
    if pastores.apagar_evento_lider(evento_id):
        flash("Evento removido do calendário.", "ok")
    else:
        flash("Evento não encontrado.", "erro")
    return redirect(url_for("pastores_admin", aba="calendario"))


@app.route("/comunicado")
def comunicado_obreiros():
    igreja = load_json("igreja.json")
    hoje = pastores.hoje()
    ano = int(request.args.get("ano", hoje.year))
    mes = int(request.args.get("mes", hoje.month))
    return render_template(
        "comunicado.html",
        igreja=igreja,
        escala=pastores.listar_escala(ano, mes),
        escala_destaque=pastores.obter_proxima_escala(),
        destaque=pastores.obter_destaque(),
        avisos=pastores.avisos_proximos(7),
        calendario=pastores.calendario_mes(ano, mes),
        ano=ano,
        mes=mes,
    )


@app.route("/calendario-lideres")
def calendario_lideres():
    igreja = load_json("igreja.json")
    hoje = pastores.hoje()
    ano = int(request.args.get("ano", hoje.year))
    mes = int(request.args.get("mes", hoje.month))
    return render_template(
        "calendario_lideres.html",
        igreja=igreja,
        calendario=pastores.calendario_mes(ano, mes),
        eventos=pastores.listar_eventos_lideres(incluir_passados=True),
        avisos=pastores.avisos_proximos(14),
        ano=ano,
        mes=mes,
    )


# ---------- Arraiá Gospel / Cantina (Diac. Cássia) ----------

@app.route("/arraial/login", methods=["GET", "POST"])
def arraial_login():
    igreja = load_json("igreja.json")
    erro = None
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if check_password_hash(ARRAIAL_PASSWORD_HASH, senha):
            session["arraial_ok"] = True
            destino = request.args.get("next") or url_for("arraial_admin")
            return redirect(destino)
        erro = "Senha incorreta. Tente novamente."
    if session.get("arraial_ok"):
        return redirect(url_for("arraial_admin"))
    info = arraial.info_evento()
    return render_template(
        "arraial_login.html",
        igreja=igreja,
        erro=erro,
        h1_responsavel=info["h1"],
        info=info,
    )


@app.route("/arraial/logout")
def arraial_logout():
    session.pop("arraial_ok", None)
    return redirect(url_for("eventos_page"))


@app.route("/arraial/admin", methods=["GET", "POST"])
@arraial_login_required
def arraial_admin():
    igreja = load_json("igreja.json")

    if request.method == "POST":
        acao = request.form.get("acao", "cantina").strip()
        if acao == "evento":
            _processar_evento_responsavel("arraial")
            return redirect(url_for("arraial_admin"))

        if acao == "convite":
            flyer_nome = None
            arquivo = request.files.get("flyer")
            if arquivo and arquivo.filename:
                ext = Path(arquivo.filename).suffix.lower()
                if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
                    flash("Flyer: use JPG, PNG, WEBP ou GIF.", "erro")
                    return redirect(url_for("arraial_admin"))
                nome_seguro = secure_filename(arquivo.filename)
                ext = Path(nome_seguro).suffix.lower() or ext
                flyer_nome = f"flyer-{uuid.uuid4().hex}{ext}"
                arquivo.save(arraial.UPLOAD_DIR / flyer_nome)

            arraial.salvar_convite(
                h1=request.form.get("h1", ""),
                convite=request.form.get("convite", ""),
                data_iso=request.form.get("data", ""),
                horario=request.form.get("horario", ""),
                local=request.form.get("local", ""),
                adulto=request.form.get("adulto", ""),
                crianca=request.form.get("crianca", ""),
                pedido=request.form.get("pedido", ""),
                flyer_arquivo=flyer_nome,
            )
            flash("Convite do Arraiá atualizado (textos e imagem).", "ok")
            return redirect(url_for("arraial_admin"))

        arraial.salvar_cantina(request.form.get("cantina", ""))
        flash("Cardápio / avisos da cantina atualizados.", "ok")
        return redirect(url_for("arraial_admin"))

    editar_evento = None
    if request.args.get("editar_evento"):
        editar_evento = pastores.obter_evento_lider(
            request.args.get("editar_evento", type=int),
            origem="arraial",
        )

    info = arraial.info_evento()
    return render_template(
        "arraial_admin.html",
        igreja=igreja,
        h1_responsavel=info["h1"],
        cantina=arraial.obter_cantina(),
        info=info,
        eventos_calendario=pastores.listar_eventos_lideres(origem="arraial"),
        editar_evento=editar_evento,
        info_evento=pastores.RESPONSAVEIS_EVENTO["arraial"],
    )


@app.route("/arraial/admin/evento/<int:evento_id>/apagar", methods=["POST"])
@arraial_login_required
def arraial_apagar_evento(evento_id: int):
    if pastores.apagar_evento_lider(evento_id, origem="arraial"):
        flash("Evento removido do calendário.", "ok")
    else:
        flash("Evento não encontrado.", "erro")
    return redirect(url_for("arraial_admin"))


@app.route("/arraial")
def arraial_page():
    igreja = load_json("igreja.json")
    info = arraial.info_evento()
    return render_template(
        "arraial.html",
        igreja=igreja,
        h1_responsavel=info["h1"],
        info=info,
        cantina=arraial.obter_cantina(),
    )


# ---------- Líderes da Mocidade ----------

@app.route("/mocidade/login", methods=["GET", "POST"])
def mocidade_login():
    igreja = load_json("igreja.json")
    erro = None
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if check_password_hash(MOCIDADE_PASSWORD_HASH, senha):
            session["mocidade_ok"] = True
            destino = request.args.get("next") or url_for("mocidade_admin")
            return redirect(destino)
        erro = "Senha incorreta. Tente novamente."
    if session.get("mocidade_ok"):
        return redirect(url_for("mocidade_admin"))
    return render_template(
        "mocidade_login.html",
        igreja=igreja,
        erro=erro,
        h1_responsaveis=mocidade.H1_RESPONSAVEIS,
        foto_lideres=mocidade.FOTO_LIDERES,
    )


@app.route("/mocidade/logout")
def mocidade_logout():
    session.pop("mocidade_ok", None)
    return redirect(url_for("eventos_page"))


@app.route("/mocidade/admin", methods=["GET", "POST"])
@mocidade_login_required
def mocidade_admin():
    igreja = load_json("igreja.json")

    if request.method == "POST":
        acao = request.form.get("acao", "post").strip()
        if acao == "evento":
            _processar_evento_responsavel("mocidade")
            return redirect(url_for("mocidade_admin"))

        titulo = request.form.get("titulo", "").strip() or "Culto dos jovens"
        texto = request.form.get("texto", "").strip()
        arquivo = request.files.get("foto")
        if not texto:
            flash("Escreva a mensagem do post.", "erro")
            return redirect(url_for("mocidade_admin"))
        if not arquivo or not arquivo.filename:
            flash("Envie uma foto para o post.", "erro")
            return redirect(url_for("mocidade_admin"))
        if not mocidade.extensao_ok(arquivo.filename):
            flash("Foto: use JPG, PNG, WEBP ou GIF.", "erro")
            return redirect(url_for("mocidade_admin"))

        nome_seguro = secure_filename(arquivo.filename)
        extensao = Path(nome_seguro).suffix.lower()
        nome_final = f"{uuid.uuid4().hex}{extensao}"
        arquivo.save(mocidade.UPLOAD_DIR / nome_final)
        mocidade.criar_post(titulo=titulo, texto=texto, foto=nome_final)
        flash("Post publicado na página inicial!", "ok")
        return redirect(url_for("mocidade_admin"))

    editar_evento = None
    if request.args.get("editar_evento"):
        editar_evento = pastores.obter_evento_lider(
            request.args.get("editar_evento", type=int),
            origem="mocidade",
        )

    return render_template(
        "mocidade_admin.html",
        igreja=igreja,
        h1_responsaveis=mocidade.H1_RESPONSAVEIS,
        foto_lideres=mocidade.FOTO_LIDERES,
        posts=mocidade.listar_posts(),
        post_ativo=mocidade.obter_post_ativo(),
        eventos_calendario=pastores.listar_eventos_lideres(origem="mocidade"),
        editar_evento=editar_evento,
        info_evento=pastores.RESPONSAVEIS_EVENTO["mocidade"],
    )


@app.route("/mocidade/admin/post/<int:post_id>/apagar", methods=["POST"])
@mocidade_login_required
def mocidade_apagar_post(post_id: int):
    if mocidade.apagar_post(post_id):
        flash("Post apagado.", "ok")
    else:
        flash("Post não encontrado.", "erro")
    return redirect(url_for("mocidade_admin"))


@app.route("/mocidade/admin/post/<int:post_id>/desativar", methods=["POST"])
@mocidade_login_required
def mocidade_desativar_post(post_id: int):
    if mocidade.desativar_post(post_id):
        flash("Post removido da página inicial.", "ok")
    else:
        flash("Post não encontrado.", "erro")
    return redirect(url_for("mocidade_admin"))


@app.route("/mocidade/admin/evento/<int:evento_id>/apagar", methods=["POST"])
@mocidade_login_required
def mocidade_apagar_evento(evento_id: int):
    if pastores.apagar_evento_lider(evento_id, origem="mocidade"):
        flash("Evento removido do calendário.", "ok")
    else:
        flash("Evento não encontrado.", "erro")
    return redirect(url_for("mocidade_admin"))


@app.route("/mocidade")
def mocidade_page():
    igreja = load_json("igreja.json")
    return render_template(
        "mocidade.html",
        igreja=igreja,
        h1_responsaveis=mocidade.H1_RESPONSAVEIS,
        foto_lideres=mocidade.FOTO_LIDERES,
        post_ativo=mocidade.obter_post_ativo(),
    )


# ---------- API ----------

@app.route("/api/info")
def api_info():
    return jsonify(load_json("igreja.json"))


@app.route("/api/cultos")
def api_cultos():
    return jsonify(load_json("cultos.json"))


@app.route("/api/eventos")
def api_eventos():
    return jsonify(load_json("eventos.json"))


@app.route("/api/contato")
def api_contato():
    igreja = load_json("igreja.json")
    return jsonify(
        {
            "nome": igreja.get("nome"),
            "endereco": igreja.get("endereco"),
            "contato": igreja.get("contato"),
        }
    )


@app.route("/api/galeria")
def api_galeria():
    posts = gallery.listar_posts_ativos()
    return jsonify({"posts": posts})


@app.route("/api/batismo")
def api_batismo():
    return jsonify(
        {
            "fotos": batismo.listar_fotos(),
            "total_inscricoes": len(batismo.listar_inscricoes()),
        }
    )


@app.route("/api/casais")
def api_casais():
    return jsonify(
        {
            "fotos": casais.listar_fotos(),
            "programacao": casais.obter_programacao(),
            "total_inscricoes": len(casais.listar_inscricoes()),
        }
    )


@app.route("/api/obreiros")
def api_obreiros():
    return jsonify(
        {
            "escala_destaque": pastores.obter_proxima_escala(),
            "destaque_culto": pastores.obter_destaque(),
            "avisos": pastores.avisos_proximos(7),
        }
    )


@app.route("/data/<path:filename>")
def data_files(filename: str):
    """Expõe os JSON para o front estático / GitHub Pages."""
    return send_from_directory(DATA_DIR, filename)


# Garante pastas/banco mesmo com Gunicorn (produção)
gallery.init_db()
batismo.init_db()
casais.init_db()
pastores.init_db()
arraial.init_db()
mocidade.init_db()


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
